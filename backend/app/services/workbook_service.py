import json
import os
import tempfile

from openpyxl import Workbook as ExcelWorkbook

from app.services.storage_service import upload_workbook
from app.models.financial_statements import FinancialStatement
from app.models.financial_analysis import FinancialAnalysis
from app.models.workbooks import Workbook
from app.models.document import Document

def generate_workbook(
    db,
    document_id
    ):

    workbook = ExcelWorkbook()

    statements = db.query(
    FinancialStatement
    ).filter(
        FinancialStatement.document_id == document_id
    ).all()

    analysis = db.query(
        FinancialAnalysis
    ).filter(
        FinancialAnalysis.document_id == document_id
    ).first()

    # Human-readable sheet names for each statement_type.
    _SHEET_NAMES = {
        # New prefixed types
        "standalone_income_statement": "Standalone Income Statement",
        "standalone_balance_sheet": "Standalone Balance Sheet",
        "standalone_cash_flow": "Standalone Cash Flow",
        "consolidated_income_statement": "Consolidated Income Statement",
        "consolidated_balance_sheet": "Consolidated Balance Sheet",
        "consolidated_cash_flow": "Consolidated Cash Flow",
        # Legacy types (backward compatibility)
        "income_statement": "Income Statement",
        "balance_sheet": "Balance Sheet",
        "cash_flow": "Cash Flow",
    }

    # Build sheets dynamically based on which statement_types exist.
    sheet_map = {}
    sheet_rows = {}
    first_sheet = True

    for statement in statements:
        st = statement.statement_type
        if st not in sheet_map:
            sheet_name = _SHEET_NAMES.get(st, st)
            if first_sheet:
                sheet = workbook.active
                sheet.title = sheet_name
                first_sheet = False
            else:
                sheet = workbook.create_sheet(sheet_name)
            sheet_map[st] = sheet
            sheet_rows[sheet] = 1

    for statement in statements:

        sheet = sheet_map.get(
            statement.statement_type
        )

        if not sheet:
            continue

        data = json.loads(
            statement.data_json
        )

        row = sheet_rows[sheet]

        for key, value in data.items():

            if isinstance(value, dict):

                sheet.cell(
                    row=row,
                    column=1,
                    value=key.upper()
                )

                row += 1

                for sub_key, sub_value in value.items():

                    sheet.cell(
                        row=row,
                        column=1,
                        value=sub_key
                    )

                    sheet.cell(
                        row=row,
                        column=2,
                        value=sub_value
                    )

                    row += 1

            else:

                sheet.cell(
                    row=row,
                    column=1,
                    value=key
                )

                sheet.cell(
                    row=row,
                    column=2,
                    value=value
                )

                row += 1
                
        sheet_rows[sheet] = row
    
    # Create the Analysis sheet after all statement sheets.
    if first_sheet:
        # No statements existed — use the default active sheet.
        analysis_sheet = workbook.active
        analysis_sheet.title = "Analysis"
    else:
        analysis_sheet = workbook.create_sheet("Analysis")

    analysis_data = json.loads(
        analysis.analysis_json
    )

    row = 1

    for key, value in analysis_data.items():

        if isinstance(value, dict):

            analysis_sheet.cell(
                row=row,
                column=1,
                value=key.upper()
            )

            row += 1

            for sub_key, sub_value in value.items():

                analysis_sheet.cell(
                    row=row,
                    column=1,
                    value=sub_key
                )

                analysis_sheet.cell(
                    row=row,
                    column=2,
                    value=sub_value
                )

                row += 1

        else:

            analysis_sheet.cell(
                row=row,
                column=1,
                value=key
            )

            analysis_sheet.cell(
                row=row,
                column=2,
                value=value
            )

            row += 1
    
    with tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsx"
        ) as temp_file:

        temp_path = temp_file.name

    workbook.save(temp_path)

    with open(temp_path, "rb") as workbook_file:

        workbook_content = workbook_file.read()

    workbook_url = upload_workbook(
        f"workbook_{document_id}.xlsx",
        workbook_content
    )

    os.remove(temp_path)

    workbook_record = Workbook(
        document_id=document_id,
        file_path=workbook_url
    )

    db.add(workbook_record)

    document = db.query(Document).filter_by(
        id=document_id
    ).first()

    if document:
        document.status = "workbook"

    db.commit()

    return {
        "message": "Workbook Generated",
        "file_path": workbook_url
    }